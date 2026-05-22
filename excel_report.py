from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

from freelance_starter.paths import ensure_parent, project_path

REQUIRED_COLUMNS = {
    "order_id",
    "date",
    "customer",
    "product",
    "category",
    "quantity",
    "unit_price",
}


def load_order_tables(input_dir: Path) -> pd.DataFrame:
    paths = sorted(
        path
        for path in input_dir.iterdir()
        if path.suffix.lower() in {".csv", ".xlsx", ".xls"}
    )
    if not paths:
        raise ValueError(f"No CSV or Excel files found in {input_dir}")

    tables: list[pd.DataFrame] = []
    for path in paths:
        if path.suffix.lower() == ".csv":
            table = pd.read_csv(path)
        else:
            table = pd.read_excel(path)
        table["source_file"] = path.name
        tables.append(table)

    return pd.concat(tables, ignore_index=True)


def clean_orders(raw_orders: pd.DataFrame) -> pd.DataFrame:
    orders = raw_orders.copy()
    orders.columns = [
        str(column).strip().lower().replace(" ", "_") for column in orders.columns
    ]

    missing = REQUIRED_COLUMNS - set(orders.columns)
    if missing:
        missing_text = ", ".join(sorted(missing))
        raise ValueError(f"Missing required columns: {missing_text}")

    text_columns = ["order_id", "customer", "product", "category"]
    for column in text_columns:
        orders[column] = orders[column].astype(str).str.strip()

    orders["date"] = pd.to_datetime(orders["date"], errors="coerce")
    orders["quantity"] = pd.to_numeric(orders["quantity"], errors="coerce")
    orders["unit_price"] = pd.to_numeric(orders["unit_price"], errors="coerce")

    orders = orders.dropna(subset=["order_id", "date", "quantity", "unit_price"])
    orders = orders[orders["quantity"] > 0]
    orders = orders[orders["unit_price"] >= 0]
    orders = orders.drop_duplicates(subset=["order_id"], keep="last")

    orders["revenue"] = (orders["quantity"] * orders["unit_price"]).round(2)
    orders["month"] = orders["date"].dt.to_period("M").astype(str)
    orders = orders.sort_values(["date", "order_id"]).reset_index(drop=True)

    ordered_columns = [
        "order_id",
        "date",
        "month",
        "customer",
        "product",
        "category",
        "quantity",
        "unit_price",
        "revenue",
        "source_file",
    ]
    existing_columns = [column for column in ordered_columns if column in orders.columns]
    return orders[existing_columns]


def summarize_by_category(cleaned_orders: pd.DataFrame) -> pd.DataFrame:
    summary = (
        cleaned_orders.groupby("category", as_index=False)
        .agg(quantity=("quantity", "sum"), revenue=("revenue", "sum"))
        .sort_values("revenue", ascending=False)
    )
    summary["revenue"] = summary["revenue"].round(2)
    return summary


def summarize_by_month(cleaned_orders: pd.DataFrame) -> pd.DataFrame:
    monthly = (
        cleaned_orders.groupby("month", as_index=False)
        .agg(orders=("order_id", "count"), quantity=("quantity", "sum"), revenue=("revenue", "sum"))
        .sort_values("month")
    )
    monthly["revenue"] = monthly["revenue"].round(2)
    return monthly


def write_excel_report(
    cleaned_orders: pd.DataFrame,
    category_summary: pd.DataFrame,
    monthly_summary: pd.DataFrame,
    output_path: Path,
) -> None:
    ensure_parent(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        cleaned_orders.to_excel(writer, sheet_name="cleaned_orders", index=False)
        category_summary.to_excel(writer, sheet_name="summary_by_category", index=False)
        monthly_summary.to_excel(writer, sheet_name="monthly_trend", index=False)

    add_category_chart(output_path)


def add_category_chart(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    worksheet = workbook["summary_by_category"]

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    if worksheet.max_row >= 2:
        chart = BarChart()
        chart.title = "Revenue by Category"
        chart.y_axis.title = "Revenue"
        chart.x_axis.title = "Category"

        data = Reference(worksheet, min_col=3, min_row=1, max_row=worksheet.max_row)
        categories = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 14
        worksheet.add_chart(chart, "E2")

    workbook.save(workbook_path)


def build_report(input_dir: Path, output_path: Path, summary_csv_path: Path | None) -> dict[str, int]:
    raw_orders = load_order_tables(input_dir)
    cleaned_orders = clean_orders(raw_orders)
    category_summary = summarize_by_category(cleaned_orders)
    monthly_summary = summarize_by_month(cleaned_orders)
    write_excel_report(cleaned_orders, category_summary, monthly_summary, output_path)

    if summary_csv_path is not None:
        ensure_parent(summary_csv_path)
        category_summary.to_csv(summary_csv_path, index=False, encoding="utf-8-sig")

    return {
        "source_rows": len(raw_orders),
        "clean_rows": len(cleaned_orders),
        "categories": len(category_summary),
    }


def parse_args() -> argparse.Namespace:
    default_input = project_path("projects", "01_excel_report_automation", "input")
    default_output = project_path(
        "projects", "01_excel_report_automation", "output", "sales_report.xlsx"
    )
    default_summary = project_path(
        "projects", "01_excel_report_automation", "output", "category_summary.csv"
    )

    parser = argparse.ArgumentParser(
        description="Merge sales files, clean data, and build an Excel report."
    )
    parser.add_argument("--input-dir", type=Path, default=default_input)
    parser.add_argument("--output", type=Path, default=default_output)
    parser.add_argument("--summary-csv", type=Path, default=default_summary)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    stats = build_report(args.input_dir, args.output, args.summary_csv)
    print(
        "Report created: "
        f"{args.output} | source rows={stats['source_rows']} "
        f"clean rows={stats['clean_rows']} categories={stats['categories']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
